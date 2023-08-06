
from calendar import c
import pandas as pd
import numpy as np
import datetime
from pathlib import Path
from openpyxl import load_workbook
import chardet

def settings_importer(file_path: Path, sheet_index:int)->list:
    '''
    returns following in a list:
    - import_cols as list
    - import_dtype as dict
    - col_rename as dic
    - date_cols as list
    - amt_cols as list
    - int_cols as list

    '''

    df_settings_file = pd.read_excel(file_path, sheet_name=sheet_index, engine='openpyxl')
    columns_in_setting = df_settings_file.columns
    check = all(item in ['import_cols', 'import_dtype', 'rename', 'convert_int', 'convert_float64', 'convert_date'] for item in columns_in_setting)

    if check:

        # picking rows till null value
        for ind, row in df_settings_file.iterrows():
            if pd.isna(row['import_cols']):
                break
        df_settings_file = df_settings_file[:ind]

        df_settings_file = df_settings_file.loc[~df_settings_file['import_cols'].isnull()]

        # list of columns to be imported

        import_cols = df_settings_file.loc[~df_settings_file['import_cols'].isnull()]['import_cols'].tolist()

        # importing column data types

        import_dtype = df_settings_file.loc[:,['import_cols','import_dtype']]
        import_dtype['import_dtype'] = import_dtype['import_dtype'].fillna('')
        import_dtype_dict = {}
        for ind, row in import_dtype.iterrows():
            if row['import_dtype']=='':
                import_dtype_dict[row['import_cols']] = 'str'
            else:
                import_dtype_dict[row['import_cols']] = row['import_dtype']

        # importing column rename dict

        df_col_rename = df_settings_file.loc[~df_settings_file['rename'].isnull(),['import_cols','rename']]
        col_rename_dict = {}
        for ind, row in df_col_rename.iterrows():
            col_rename_dict[row['import_cols']] = row['rename']

        # list of cols to be converted to date, amount, int

        date_cols = df_settings_file.loc[~df_settings_file['convert_date'].isnull()]['import_cols'].tolist()
        amt_cols = df_settings_file.loc[~df_settings_file['convert_float64'].isnull()]['import_cols'].tolist()
        int_cols = df_settings_file.loc[~df_settings_file['convert_int'].isnull()]['import_cols'].tolist()

        return {'import_cols':import_cols,
            'import_dtype':import_dtype_dict,
            'col_rename':col_rename_dict,
            'date_cols':date_cols,
            'amt_cols':amt_cols,
            'int_cols':int_cols}

    else:
        return False

def col_dtype_converter(df:pd.DataFrame, date_cols:list=False, day_first = False, amt_cols:list=False, code_cols:list=False) -> pd.DataFrame:
    '''
    converts date, amount and int columns from string to these dtypes. Date column can handle dates with - or /
    Removes -, NA, na
    '''

    if date_cols:
        for col in date_cols:
            if df[col].dtype == 'datetime64[ns]':
                pass
            else:
                print(f'Attempting to convert {col} to date!!')
                df[col] = df[col].fillna('')
                df[col] = df[col].replace('-','').replace('NA','').replace('na','')
                df[col] = df[col].astype(str).str.replace('-','/')
                df[col] = pd.to_datetime(df[col], errors='raise', dayfirst=day_first)

    if amt_cols:
        for col in amt_cols:
            if df[col].dtype == 'float64':
                pass
            else:
                print(f'Attempting to convert {col} to amount!!')
                df[col] = df[col].fillna('0')
                df[col] = df[col].replace('-', '0').replace('NA','0').replace('na','0')
                df[col] = df[col].str.replace(',','').replace(')','').replace('(','-').astype(float)

    if code_cols:
        for col in code_cols:
            if df[col].dtype == 'float64':
                print(f'Attenpting to convert {col} from {df[col].dtype} into int')
                df[col] = df[col].replace('-',0).replace('NA',0).replace('na',0)
                df[col].fillna(0, inplace=True)
                df[col] = df[col].astype(np.int64)
            elif df[col].dtype == 'object':
                print(f'Attenpting to convert {col} from {df[col].dtype} into int')
                df[col] = df[col].fillna('0')
                df[col] = df[col].replace('-','0').replace('NA','0').replace('na','0')
                df[col] = df[col].astype(np.int64)

    return df

def excel_writer(df:pd.DataFrame, output_file:Path, sheet_name):
    '''
    Checks if file / sheet already exists. If yes, appends data to existing file / sheet else creates new
    file / sheet.
    '''
    if output_file.exists():
        # checking last used row before appending
        wb = load_workbook(output_file, read_only=True)
        sheets = wb.sheetnames
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            last_row = len(list(ws.rows))
        else:
            last_row=0
        wb.close()

        # appending to file in existing sheet
        with pd.ExcelWriter(output_file, mode='a', date_format='DD/MM/YYYY', engine='openpyxl', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=last_row, header=False)
    else:
        # writing to a new file if it does not exist
        with pd.ExcelWriter(output_file, mode='w', date_format='DD/MM/YYYY', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

def df_col_search(df:pd.DataFrame, srch_string: str):
    '''
    Searches given string in column header in lower case and returns list of columns where string is found.
    Case insensitive.
    '''
    col_list = []
    for col in df.columns:
        if srch_string in col.lower():
            col_list.append(col)
    return col_list

def dtype_to_df(df: pd.DataFrame)->None:
    '''
    Accepts a dataframe and returns df containing column dtypes
    '''
    df_cols = pd.DataFrame(columns=['Type'], data=df.dtypes)
    df_cols.index.name = 'Col'
    df_cols.reset_index(inplace=True)
    return df_cols

def ageing_creator(df:pd.DataFrame, col:str, to_date:datetime.date=datetime.date.today(), bins:list=[0,30,60,90,5000], bins_labels:list=['0 to 30', '30 to 60','60 to 90', 'above 90'])->pd.Series:
    '''Creates following parameters and eturns dataframe with ageing:
    df = pandas dataframe
    col = column name of the column in dataframe. This should contain data in date format.
    bins = list of integers containing the bins
    bins_labels = labels to be used for bins    

    Returns binned series which can be directly added to the source dataframe
    '''

    if len(bins) != len(bins_labels)+1:
        return False

    diff = to_date - df[col]

    bins_converted = []
    for bin in bins:
        bins_converted.append(datetime.timedelta(bin))

    binned = pd.cut(diff, bins=bins_converted, include_lowest=True, right=True, labels=bins_labels)
    del diff

    return binned

def csv_encoding(file_path:Path)->list:
    '''
    Takes path to a csv file and determines the csv encoding. Returns [encoding, confidence, language]
    '''
    if file_path.is_file and file_path.suffix == '.csv':
        with open(file_path, 'rb') as rawdata:
            result = chardet.detect(rawdata.read(10000))
        return result
    else:
        return [None, 'Path is not a file or a csv']

def df_header_cleaner(df:pd.DataFrame, na_cols:list=False)->pd.DataFrame:
    '''
    Cleans blank rows. Sets first non blank row as header. Also removes rows with null values in na_cols.
    '''
    df = df.dropna(how='all')
    df.columns = df.iloc[0].values
    df = df[1:]
    df = df.reset_index(drop=True, inplace=False)
    if na_cols!=False:
        df = df.dropna(how='any', subset=['Sl No'], inplace=True)
    return df