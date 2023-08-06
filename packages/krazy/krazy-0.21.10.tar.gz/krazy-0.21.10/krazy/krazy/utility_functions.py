#List of useful pandas dataframe commands

import pandas as pd
from openpyxl import load_workbook
import glob
import re
from io import StringIO
import os
import datetime
import csv
import pytesseract
from PIL import Image
import pytz
import time
from PyPDF2 import PdfWriter, PdfReader, PdfMerger
from pathlib import Path
import sys

def df_retainCols(dfdata, collist): #deletes columns in df which are not there in list of columns
    # collist (a list) should contain the list of columns to be retained
    dfdata.drop(columns=[col for col in dfdata if col not in collist], inplace=True)
    return dfdata

def df_delCcols(dfdata, collist): #deletes columns in df which are there in list of columns
    # collist (a list) should contain the list of columns to be deleted
    dfdata.drop(columns=[col for col in dfdata if col in collist], inplace=True)
    return dfdata

def df_renameCols(dfdata, coldict): #renames columns based on dictionery of old vs new column names
    # coldict (a dictionery) should contain list of column name and names to be replaced with
    dfdata.rename(columns=coldict,inplace=True)
    return dfdata

def list_sort(listToSort, direction): #sorts list (not dataframe)
    '''
    Sorts list. Pass anything non blank in direction to sort in reverse order.
    '''
    if direction == '':
        listToSort.sort()
    else:
        listToSort.sort(reverse=True)
    return listToSort

def df_change_index(dfdata, colList):
    dfdata.set_index(colList, drop=True, inplace=True) #change index col
    return dfdata

def df_sort(dfdata,colList):
    if not colList:
        dfdata.sort_index(ascending=True,inplace=True) #sort on index column
    else:
        df.sort_values(by=colList,ascending=True,inplace=True) #sort on column value

def df_find_index(dfdata,col_name):
    colindex = dfdata.columns.get_loc(col_name) #finds index number of col
    return colindex

def df_seq(dfdata, colseq):
    dfdata = dfdata.reindex(columns=colseq) #changes sequence of columns

#example of creating dataframe using dictionery
def data_for_df():
    data = {
    "Duration":{
        "0":60,
        "1":60,
        "2":60,
        "3":45,
        "4":45,
        "5":60
    },
    "Pulse":{
        "0":110,
        "1":117,
        "2":103,
        "3":109,
        "4":117,
        "5":102
    },
    "Maxpulse":{
        "0":130,
        "1":145,
        "2":135,
        "3":175,
        "4":148,
        "5":127
    },
    "Calories":{
        "0":409,
        "1":479,
        "2":340,
        "3":282,
        "4":406,
        "5":300
    }
    }
    df = pd.DataFrame(data)
    return df

def df_import_single_excel(filepath, shtname): #import using openpyxl
    # full path & should include filename. Further, path should be with / and not \
    # sht to contain sheet name to be imported
    wb = load_workbook(filepath)
    ws = wb[shtname]
    data = ws.values
    col = next(data)
    dfdata = pd.DataFrame(data,columns=col)
    return dfdata

def df_multi_excel_import(folderpath, shtname, srch, excl):
    #imports sheet from excel using openpyxl
    ''' Requires following variables
    folderpath i.e. path to folder from where files are to be imported
    shtname i.e. name of sheet in excel file from where data is to be imported
    srch i.e. string to be included in file name search
    excl i.e. string to be excluded in file name search
    '''
    filenames = glob.glob(folderpath + "*.xlsx")
    dfs = []
    for filename in filenames:
        if re.search(srch.lower(),filename.lower()):
            if re.search(excl,filename):
                continue
            else:
                wb = load_workbook(filename)
                if (shtname.strip()==''):
                    ws = wb.active
                else:
                    ws = wb[shtname]
                data = ws.values
                col = next(data)
                dfs.append(pd.DataFrame(data,columns=col))
    # Concatenate all data into one DataFrame
    dfjoined = pd.concat(dfs, ignore_index=True)
    return dfjoined

def df_multi_csv_importer(fpath, srch, excl, types):
    #imports BR files from a path
    filenames = glob.glob(fpath + "*.csv")
    dfs = []
    for filename in filenames:
        if re.search(srch.lower(),filename.lower()):
            if re.search(excl.lower(),filename.lower()):
                continue
            else:
                if (types == ''):
                    dfs.append(pd.read_csv(filename,low_memory=False))
                else:
                    dfs.append(pd.read_csv(filename,dtype=types))
    # Concatenate all data into one DataFrame
    dfjoined = pd.concat(dfs, ignore_index=True)
    return dfjoined

def df_to_sql(dfdata,table,colList,connection):
    # this inserts data from dataframe into database table
    # for this to work, columns and column sequence should be exactly in same manner as in database
    cur = connection.cursor()
    imFile = StringIO()
    dfdata.to_csv(imFile, index=False)
    imFile.seek(0)
    next(imFile)
    cur.copy_from(imFile,table, sep=',',columns=colList)
    imFile.close()
    connection.commit()
    cur.close()
    connection.close()

def df_compare(df1, df2, whichStr, colStr):
    if whichStr == '':
        comparisondf = df1.merge(df2, indicator = True, how = 'outer', on= colStr, validate = '1:1')
    else:
        comparisondf = df1.merge(df2, indicator = 'Exist', how = whichStr, left_on='InvNo', right_on='InvNo', validate = '1:1')
    return comparisondf

def df_group(dfdata, grColList, colGroupDict):
    dfdata = dfdata.groupby(grColList).agg(colGroupDict)
    return dfdata

def df_append(df1, df2):
    df1 = df1.append(df2, ignore_index=True)
    return df1

def convertpath(fpath):
    if fpath[-1] == '\\':
        fpath = fpath.replace('\\','/')
    else:
        fpath = fpath.replace('\\','/') + '/'
    return fpath

def get_file_details(file_path):
    '''Gets file details for the given file. Created time, modified time, size in MB'''

    try:
        # creation date and time
        ct = os.path.getctime(file_path)
        if ct not in [None,'']:
            ct = time.ctime(ct)
            ct = datetime.datetime.strptime(ct,'%a %b %d %H:%M:%S %Y')
            ctstr = datetime.datetime.strftime(ct,'%d/%m/%Y')
        
        # modified date and time
        mt = os.path.getmtime(file_path)
        if mt not in [None,'']:
            mt = time.ctime(mt)
            mt = datetime.datetime.strptime(mt,'%a %b %d %H:%M:%S %Y')
            mtstr = datetime.datetime.strftime(mt,'%d/%m/%Y')

        # size in MB
        sz = os.path.getsize(file_path)
        if sz not in [None,'']:
            sz = round(sz/1024/1024,2)

    except:
        ct = None
        ctstr = None
        mt = None
        mtstr = None
        sz = None

    return [ct,mt,sz]

def get_full_file_list_df(dirName:str)->pd.DataFrame:
    ''' gets full file and directory tree under given directory '''

    dir_list = []
    dir_list.append(dirName)
    dffiles = pd.DataFrame({'Path':[], 'File':[], 'Creation':[], 'Modified': [], 'Size (MB)': []})

    for dir in dir_list:
        dir_content_list = os.listdir(dir)
        for entry in dir_content_list:
            full_path = os.path.join(dir,entry)
            if os.path.isdir(full_path):
                dir_list.append(full_path)
            else:
                file_details = get_file_details(full_path)
                dffiles.loc[len(dffiles)] = [dir, entry, file_details[0], file_details[1],
                    file_details[2]]
    return dffiles

def write_To_csv(header_tpl, data_tpl, filePath):
    outFile = open(filePath)
    writer = csv.writer(outFile)
    writer.writerow(header_tpl)
    writer.writerows(data_tpl)
    outFile.close()

def read_from_csv(filePath):
    inFile = open(filePath)
    reader = csv.reader(inFile)
    dataList = []
    for row in reader:
        dataList.append(row)
    inFile.close()
    return dataList

class date_converter:
    #takes string in dd/mm/yyyy or dd/mm/yy format, checks data validity and converts into
    # proper date in final_date attribute
    # this can accept either / or - as date separator
    '''following attributes are available:
    .final_date = gives date converted into proper date format
    .validity = True / False showing valid input
    .type = '/' or '-' or '' showing date separator used in input
    .manual_mapped_vals = values which were not auto mapped but mapped based on d/m/y sequence
    .manual_mapped_keys = keys which were not auto mapped but mapped based on d/m/y sequence            

    '''
    def __init__(self,date_string):
        self.date_string = date_string.strip(' ')
        self.__is_valid()
        if self.__is_valid:
            self.__convert_date()

    def __is_valid(self):

        # initializing variables
        slash_type = None
        hyphen_type = None
        first = ''
        second = ''
        third = ''

        # checking if / is used as date separator
        slash_type = self.date_string.count('/')
        if slash_type == 2:
            num_check = len(self.date_string.replace('/',''))
            first, second, third = self.date_string.split('/')

        # checking if - is used as date separator
        hyphen_type = self.date_string.count('-')
        if hyphen_type == 2:
            num_check = len(self.date_string.replace('-',''))
            first, second, third = self.date_string.split('-')

        #creating final validity
        self.validity = False
        
        if len(first) != 0 and len(second) != 0 and len(third) != 0:
            if (slash_type == 2 or hyphen_type ==2) and (4 <= num_check <= 8):
                if len(first) < 5 and len(second) < 5 and len(third) < 5:
                    if 2 < len(first) < 5:
                        if len(second)< 3 and len(third) < 3:
                            if int(second) < 32 and int(third) < 13:
                                self.validity = True
                            elif int(third) < 32 and int(second) < 13:
                                self.validity = True
                    elif 2 < len(second) < 5 :
                        if len(first) < 3 and len(third) < 3:
                            if int(first) < 32 and int(third) < 13:
                                self.validity = True
                            if int(third) < 32 and int(first) < 13:
                                self.validity == True
                    elif 2 < len(third) < 5:
                        if len(first) < 3 and len(second) < 3:
                            if int(first) < 32 and int(second) < 13:
                                self.validity = True
                            if int(second) < 32 and int(first) < 13:
                                self.validity = True

        if self.validity == True:
            if slash_type == 2:
                self.type = '/'
            if hyphen_type == 2:
                self.type = '-'
        else:
            self.type = ''


    def __convert_date(self):
        if self.validity == True:

            check = False
            if self.type == '/':
                first, second, third = self.date_string.split('/')
            else:
                first, second, third = self.date_string.split('-')
            first = int(first)
            second = int(second)
            third = int(third)
            
            # initializing working variables
            val_unmapped = [first, second, third]
            mapping = {'year': None, 'month': None, 'day': None}
            keys_unmapped = ['day','month','year']

            # getting year with 4 digits
            if len(str(first)) == 4 or first > 31:
                mapping.update(year = first,)
                val_unmapped.remove(first)
                keys_unmapped.remove('year')
            elif len(str(second))== 4 or second > 31:
                mapping.update(year = second,)
                val_unmapped.remove(second)
                keys_unmapped.remove('year')
            elif len(str(third))== 4 or third > 31:
                mapping.update(year = third,)
                val_unmapped.remove(third)
                keys_unmapped.remove('year')

            # getting day > 31            
            if 12 < first < 31:
                if first in val_unmapped:
                    mapping.update(day = first,)
                    val_unmapped.remove(first)
                    keys_unmapped.remove('day')
            elif 12 < second < 31:
                if second in val_unmapped:
                    mapping.update(day = second,)
                    val_unmapped.remove(second)
                    keys_unmapped.remove('day')
            elif 12 < third < 31:
                if third in val_unmapped:
                    mapping.update(day = third,)
                    val_unmapped.remove(third)
                    keys_unmapped.remove('day')
            
            # mapping remaining in entered sequence
            ind = 0
            for elem in keys_unmapped:
                tmp = {elem:val_unmapped[ind]}
                ind += 1
                mapping.update(tmp)
            
            if len(str(mapping['year'])) == 3:
                mapping['year'] = int(('2' + str(mapping['year'])))
            if len(str(mapping['year'])) == 2:
                mapping['year'] = int(('20' + str(mapping['year'])))
            if len(str(mapping['year'])) == 1:
                mapping['year'] = int(('200' + str(mapping['year'])))

            # creating date in date format
            self.final_date = datetime.datetime(mapping['year'],mapping['month'],mapping['day'])

            self.manual_mapped_vals = val_unmapped
            self.manual_mapped_keys = keys_unmapped
        else:
            self.final_date = None

def number_word_converter(amount:int) -> str:
    dict1 = {'0':'','1': ' One', '2': ' Two', '3': ' Three','4':' Four','5':' Five','6':' Six',
        '7':' Seven','8':' Eight','9':' Nine','10':' Ten', '11':' Eleven','12':' Twelve','13':' Thirteen',
        '14':' Forteen','15':' Fifteen','16':' Sixteen','17':' Seventeen','18':' Eighteen',
        '19':' Nineteen','20':' Twenty'}
    dict2 = {'0':'','1':' One','2':' Twenty','3':' Thirty','4':' Forty','5':' Fifty','6':' Sixty',
        '7':' Seventy','8':' Eighty','9':' Ninety'}
    dict_place = {0:'',1:'',2:' Hundred',3:' Thousand',4:'',5:' Lac',6:'',7:' Crore',8:'',
        9:' Hundred',10:'',11:' Thousand',12:'',13:' Lac'}
    amt_str = ''
    amount = str(amount)
    al = [x for x in reversed(amount)]
    
    max_digits = len(al) - 1

    for counter in range(max_digits+1):

        if counter == 2 and int(al[counter]) != 0:
            amt_str = dict_place[counter] + amt_str
        elif counter in [3,5,7,9]:
            if counter == max_digits:
                if int(al[counter]) != 0:
                    amt_str = dict_place[counter] + amt_str
            else:
                if (int(al[counter]) + int(al[counter+1])) != 0:
                    amt_str = dict_place[counter] + amt_str
        else:
            pass

        if counter in [0,3,5,7,9]:
            if counter == max_digits:
                amt_str = dict1[al[counter]] + amt_str
            elif al[counter+1] == '1':
                pass
            else:
                amt_str = dict1[al[counter]] + amt_str
        elif counter in [1,4,6,8,10]:
            if al[counter] == '1':
                val = al[counter] + al[(counter-1)]
                amt_str = dict1[val] + amt_str
            else:
                amt_str = dict2[al[counter]] + amt_str
        else:
            amt_str = dict1[al[counter]] + amt_str
    amt_str = amt_str.strip()
    return amt_str

def payment_ocr(img_path):
    '''
    ocr on google pay and paytm screenshots
    returns following:
    1. if successfule: [success, template, paid_to, paid_on, amount, upi_id]
    2. if image file not found: [FNF]
    3. if template not found: [TNF]
    '''
    # ocr on google pay and paytm screenshots
    # returns success, template, paid_to, paid_on, amount, upi_id if successful else returns None

    paid_to = None
    paid_on = None
    amount = None
    upi_id = None
    template = None
    success = False
    tz = pytz.timezone('Asia/Kolkata')

    if os.path.exists(img_path):

        # doing OCR and initializing text
        text = pytesseract.image_to_string(Image.open(img_path))
        data = []
        text_lower = text.lower()
        data = text.split('\n')
        data_lower = text_lower.split('\n')

        #checking gpay
        gpay = text_lower.find('g pay')
        gpay2 = -1
        if gpay > 0:
            gpay2 = text.lower().index('to')
        gpayv = text_lower.find('paid to')
        
        # checking paytm
        paytm = text_lower.find('paid successfully to')
        paytm2 = text_lower.find('paytm wallet')
        
        try:

            # processing if matched with gpay template v2
            if gpay != -1 and gpay2 == 0 and gpayv == -1:
                success = True
                template = 'Gpay v2' # used by google pay in 2021

                identifier = data_lower.index('upi transaction id')

                upi_id = data[identifier + 1]
                paid_to = data[identifier - 11].lower().replace('to ','').strip().title()
                if paid_to in [None,'']:
                    paid_to = data[identifier - 10].lower().replace('to ','').strip().title()

                paid_on = data[identifier - 4]
                paid_on = paid_on.replace('© Completed + ','').strip()
                paid_on = paid_on.replace('© Completed - ','').strip()
                paid_on = paid_on.replace('© Completed « ','').strip()
                paid_on = datetime.datetime.strptime(paid_on,'%d %b, %I:%M %p')
                paid_on = datetime.datetime(datetime.date.today().year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

                amount = int(float(data[identifier - 8].replace('=','').strip().replace(',','')))
            
            # processing if matched with gpay template v1
            elif gpay != -1 and gpayv != -1:
                success = True
                template = 'Gpay v1'

                identifier = data_lower.index('upi transaction id')

                upi_id = data[identifier + 1]
                paid_to = data[identifier - 7].title()

                paid_on = data[identifier - 2]
                paid_on = datetime.datetime.strptime(paid_on,'%d %b %Y %I:%M %p')
                paid_on = datetime.datetime(datetime.date.today().year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

                scrap = data[identifier - 3].split('= ')
                amount = int(float(scrap[-1].strip()))

            # processing if matched with paytm
            elif paytm != -1:
                success = True
                template = 'Paytm'

                identifier = data_lower.index('paid successfully to')

                upi_id = data[identifier + 6].split(':')[1].strip()

                paid_to = data[identifier + 2]
            
                amount = data[identifier + 4]
                amount = amount.replace('@','')
                amount = amount.replace('&','')
                amount = amount.replace('°','')
                amount = amount.strip()
                amount = int(float(amount[1:]))

                paid_on = data[identifier + 7]
                paid_on = datetime.datetime.strptime(paid_on.upper(),'%I:%M %p, %d %b %Y')
                paid_on = datetime.datetime(paid_on.year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)
            
            elif paytm2 > -1:
                success = True
                template = 'Paytm v2' # paytm transaction screen

                identifier = data_lower.index('paytm wallet')

                upi_id = data[identifier + 2].split(':')[1].strip()

                paid_to = data[identifier - 3]
            
                amount = data[identifier - 9].strip()
                amount = int(float(amount))

                paid_on = data[identifier - 7].strip()
                paid_on = datetime.datetime.strptime(paid_on.upper(),'%d %b %Y, %I:%M %p')
                paid_on = datetime.datetime(paid_on.year, paid_on.month,paid_on.day,
                    paid_on.hour,paid_on.minute,tzinfo=tz)

            else:
                pass
            
            if success == True:
                return [success, template, paid_to, paid_on, amount, upi_id]
            else:
                return ['TNF', text]
        
        except:
            print(data)
            return ['TNF', text]

    else:
        return ['FNF']

def number_converter(str_num: str) -> int:

    if isinstance(str_num,int) or isinstance(str_num,float):
        return [True, int(str_num)]
    elif str_num in [None]:
        return [True, 0]
    elif isinstance(str_num,str):
        if str_num.replace('.','').strip().isdigit():
            str_num = int(float(str_num))
            return [True, str_num]
        elif str_num.strip() == '':
            return [True, 0]
        else:
            return [False]
    else:
        return [False]

class TimeKeeper():
    def __init__(self, supress:str) -> None:
        '''
        pass y flag to supresses printing by this class so that user can print as per his needs
        '''
        self.start_time = datetime.datetime.now()
        print(f'\nStarted at: {self.start_time}')
        self.logs = []
        self.logs.append(self.start_time)
        self.supress = supress
    
    def log(self) -> datetime:
        '''
        returns difference between last log and 2nd last log entry
        '''
        self.logs.append(datetime.datetime.now())
        diff = self.logs[-1] - self.logs[-2]
        if self.supress == 'y':
            pass
        else:
            print(f'\nLap runtime (sec): {diff}')
            print(f'Total runtime (sec): {self.logs[-1] - self.start_time}')
        
        return diff

    def get_log(self, log_idx:int) -> datetime:
        '''
        returns difference between given log and start time
        '''
        diff = self.logs[log_idx], self.start_time
        if self.supress == 'y':
            pass
        else:
            print(f'\nRuntime since start (sec): {diff}')
        
        return diff

    def end(self) -> datetime:
        self.end_time = datetime.datetime.now()
        self.logs.append(self.end_time)
        diff = self.end_time - self.start_time
        if self.supress == 'y':
            pass
        else:
            print(f'\nTotal runtime (sec): {diff}')
            print(f'End time: {self.end_time}')
        
        return diff

def progress_bar(iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = '\r'):
    """
    Call in a loop to create terminal progress bar
    @params:
        iteration   - Required  : current iteration (Int)
        total       - Required  : total iterations (Int)
        prefix      - Optional  : prefix string (Str)
        suffix      - Optional  : suffix string (Str)
        decimals    - Optional  : positive number of decimals in percent complete (Int)
        length      - Optional  : character length of bar (Int)
        fill        - Optional  : bar fill character (Str)
        printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
    """

    percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
    filledLength = int(length * iteration // total)
    bar = fill * filledLength + '-' * (length - filledLength)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
    # Print New Line on Complete
    if iteration == total: 
        print()

def check_path(path: str, make_dir=False)->list:
    '''
    Accepts a path, check if it exists. Creates directory if make_dir = True.
    Returns a list as follows:
    [True] - it exists and is a directory
    [True, 'Directory Created'] - it does not exists but make_dir = True
    [False, 'File] - it exists but is a file
    ['False', 'Does not exist'] - if it does not exist and make_dir is false
    ['False', err] in case there is an error in code

    '''
    try:

        path_p = Path(path)
        if path_p.exists():
            if path_p.is_dir():
                return [True]
            else:
                return [False, 'File']
        else:
            if make_dir:
                path_p.mkdir()
                return [True, 'Directory Created']
            else:
                return [False, 'Does not exist']
    except Exception as err:
        return [False, err]

def pdf_page_extractor(source_path: str, dest_path: str, no_of_pages:int, auto_save_output=False)->list:
    '''
    accepts path to a folder, extracts number of pages from each pdf file in that folder.
    Returns list as given below:
    [True, pdf file] - if all goes fine. pdf file is one consolidated file with given number of pages from each file
    [False] - if there is an error
    '''
    try:
        check_1, check_2 = False, False

        # convert path to Path object
        source_path_p=Path(source_path)    
        dest_path_p=Path(dest_path)

        # check if input exists
        if check_path(source_path_p):
            check_1=True
        else:
            print('Source path is not a valid directory. Give a valid directory!')
            return [False]

        if check_path(dest_path_p, make_dir=True):
            check_2=True


        if check_1 and check_2:

            # get list of files in source folder
            list_of_files = os.listdir(source_path_p)

            # create empty pdf file in memory
            out_pdf_file = PdfWriter()

            # iterate over files
            for file in list_of_files:
                # create full path
                full_path = source_path_p.joinpath(file)
                # check if file is pdf
                if full_path.suffix == '.pdf':
                    # read file
                    pdf_file = PdfReader(full_path)
                    # get total pages
                    total_pages = len(pdf_file.pages)
                    if total_pages ==0:
                        pass
                    else:
                        # extract and add pages
                        for page in range(0,min(no_of_pages, total_pages)):
                            out_pdf_file.add_page(pdf_file.pages[page])
                
            
            if auto_save_output:
                # save output pdf in destination directory

                # check if output file has atleast 1 page
                out_pages = len(out_pdf_file.pages)
                if out_pages>0:
                    with open(dest_path_p.joinpath('Output.pdf'), 'wb') as out_stream:
                        out_pdf_file.write(out_stream)
                
            return [True, out_pdf_file]
            
    except Exception as err:
        return [False, err]

def week_to_date(year:int, week:str, day_to_extract)->datetime.date:
    '''
    Accepts year as integer, week as string in pattern W1, day_to_extract in pattern 0 to 6 or "all". All extracts all dates in a week and returns a list.
    '''
    if type(day_to_extract)==str:
        if week[0]=='W' and week[1:].isdigit():
            if day_to_extract=='all':
                dates = []
                for i in range(1,8):
                    if i < 7:
                        day_extracted = datetime.datetime.strptime(str(year) + '-' + week + f'-{i}', '%Y-W%W-%w')
                        dates.append(day_extracted)
                    elif i == 7:
                        day_extracted = datetime.datetime.strptime(str(year) + '-' + week + '-0', '%Y-W%W-%w')
                        dates.append(day_extracted)
                return dates
            else:
                return [False, 'Invalid String']

    elif type(day_to_extract)==int:
        if 0 <= day_to_extract <= 6:
            return datetime.datetime.strptime(str(year) + '-' + week + f'-{day_to_extract}', '%Y-W%W-%w')
        else:
            return [False, 'Invalid integer. It can only be between 0 to 6.']
    
    else:
        return [False, 'Invalid parameter type. Either give "all" or an integer between 0 to 6.']