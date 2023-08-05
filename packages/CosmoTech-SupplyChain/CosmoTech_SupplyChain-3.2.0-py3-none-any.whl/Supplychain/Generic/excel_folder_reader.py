import glob
import io
import json
import os

from openpyxl import load_workbook

from Supplychain.Generic.folder_io import FolderReader


class ExcelReader(FolderReader):

    def refresh(self):
        filename = glob.glob(os.path.join(self.input_folder, "*.xlsx"))[0]
        self.files = dict()
        wb = load_workbook(filename, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self.files[sheet_name] = list()
            headers = next(sheet.iter_rows(max_row=1, values_only=True))

            def item(_row: tuple) -> dict:
                d = {k: v for k, v in zip(headers, _row) if k is not None}
                return d if any(v is not None for v in d.values()) else {}

            for r in sheet.iter_rows(min_row=2, values_only=True):
                row = item(r)
                if not row:
                    continue
                new_row = dict()
                for key, value in row.items():
                    try:
                        # Try to convert any json row to dict object
                        converted_value = json.load(io.StringIO(value))
                    except (json.decoder.JSONDecodeError, TypeError):
                        converted_value = value
                    if converted_value is not None or self.keep_nones:
                        new_row[key] = converted_value
                if new_row:
                    self.files[sheet_name].append(new_row)

    def __init__(self,
                 input_folder: str = "Input",
                 keep_nones: bool = True):

        FolderReader.__init__(self,
                              input_folder=input_folder,
                              keep_nones=keep_nones)

        self.refresh()
