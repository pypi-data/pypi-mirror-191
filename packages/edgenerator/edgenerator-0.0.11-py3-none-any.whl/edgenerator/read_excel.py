import openpyxl as xl
from .state_classes import Pathway, State

# o=======================================================o
#                        Functions                     
# o=======================================================o

def read_excelfile(path) -> dict:
    """
    Reads an excel file and returns a dictionary with the data
    Input: 
        path: path to the excel file
    Output:
        data_dic: dictionary with the data
        
    Data structure of dictionary with example items:
    data_dic = {
        'energy_type': 'E', 'H', or 'G'     # energy type 
        'outputfile': 'some name'           # name of output file
        'profile': 'default' or 'user_name'
        'paths': [
            ::Pathway:: instance containing ::States:: instances
            ::Pathway:: instance containing ::States:: instances
            .
            .
            .
            ]        
    }
    """
    data_dic = {}
    # Load whole workbook including different sheets
    wb_obj = xl.load_workbook(path)

    # Get the active sheet (default is the first sheet)
    sheet_obj = wb_obj.active


    # General reading settings
    row_limit = 100
    column_limit = 100
    start_reading_index = -1
    
    # FIND ROW WHERE READING BEGINS THROUGH 'BEGIN' STATEMENT
    for i_row in range(1,row_limit):    
        if str(sheet_obj.cell(row=i_row, column=1).value).upper() == "BEGIN":
            start_reading_index = i_row
            break

    # SETTINGS READING BLOCK
    # This block reads the settings from the excel file
    # These entail type of energy, output file name, profile, etc.
    # The block ends with an "End" statement or when the "Paths" row is reached
    settings_block = False
    path_block = False
    headers = []
        
    i_row = start_reading_index
    j_col = 2
    while i_row < row_limit:
        # Loops over the rows and stops when it reaches the "paths" row
        col1 = sheet_obj.cell(row=i_row, column=1).value
        col2 = sheet_obj.cell(row=i_row, column=2).value
        if col1 == "end": break
        if str(col1).lower() == "settings":
            settings_block = True
        if str(col1).lower() == "paths":
            settings_block = False
            path_block = True
            while j_col < column_limit:
                # looping over the headers and stops at an "end" statement (column-wise)
                if str(sheet_obj.cell(row=i_row, column=j_col).value).lower() == "end": 
                    break
                headers.append(sheet_obj.cell(row=i_row, column=j_col).value)
                j_col += 1     
            i_row += 1
            break
        if settings_block == True and col1 != None:
            data_dic[col1] = col2        
        i_row += 1

    # PATHS READING BLOCK
    # This block reads the pathways from the excel file and stores them in a list
    # These entail the name of pathways and the states (instances of ::State::)
    # The block ends with an "End" statement (row-wise)
    if path_block == True:
        data_dic["paths"] = []
        # Loop over the rows and columns
        while i_row < row_limit: 
            if str(sheet_obj.cell(row=i_row, column=1).value) == "END": break
            pathway = Pathway()
            pathway.name = sheet_obj.cell(row=i_row, column=1).value
            for j_col in range(2,len(headers)+2):
                cell_value = sheet_obj.cell(row=i_row, column=j_col).value
                if cell_value != None:
                    state = State()
                    state.label = f"{i_row+2-len(headers)}_{headers[j_col-2]}"
                    state.energy = cell_value
                    pathway.add_state(state)
                j_col += 1
            data_dic["paths"].append(pathway)
            i_row += 1
    return data_dic