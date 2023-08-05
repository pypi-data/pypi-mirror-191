# 举例
# file = 'data/5.baidu-员工的人员信息.xls'
# excel = excel(file)
# excel.sheet_xuanze()跳转表单
# print(excel.all_datas())该表单所有数据
# print(excel.all_title())该表单所有标题
# print(excel.bufen_titles(['花名', '序号']))部分表单名
# print(excel.search_lies(['密码', '序号']))搜索标题对应的数据

import openpyxl
import xlrd

class XlsExcel:
    def __init__(self, file):

        self.sheets = xlrd.open_workbook(file).sheets()
        self.data = []
        self.sheet = self.sheets[0]

    def sheet_xuanze(self, n=0):
        self.sheet = self.sheets[n]
        return self.sheet

    def datas(self):
        datas = []
        print('该表单数据如下'.center(33, '-'))
        for i in range(1, self.sheet.nrows):
            row_data = self.sheet.row_values(rowx=i, start_colx=0, end_colx=None)
            datas.append(row_data)
            print(row_data)
        print('数据读取完毕'.center(33, '-'))
        return datas

    def all_title(self):
        return self.sheet.row_values(rowx=0, start_colx=0, end_colx=None)

    def bufen_titles(self, lie_titiles=None):
        if not lie_titiles:
            return self.all_title()
        title = []
        for i in lie_titiles:
            if i in self.all_title():
                title.extend([i])
        return title

    def count(self):
        return self.sheet.nrows - 1

    def search_lies(self, lie_titles):
        row_datas = []
        for i in range(1, self.sheet.nrows):
            row_data = []
            for lie_title in lie_titles:
                if lie_title in self.all_title():
                    row_data.extend(
                        [self.sheet.row_values(rowx=i, start_colx=0, end_colx=None)[self.search_index(lie_title)]])
                    # [i[self.search_index(lie_title)]
            row_datas.append(row_data)
        return row_datas

    def search_index(self, lie_title):
        if lie_title in self.all_title():
            return self.all_title().index(lie_title)


def excel(file):
    file_3, file_4 = str(file)[-3:], str(file)[-4:]
    if file_3 == 'xls':
        return XlsExcel(file)
    if file_4 == 'xlsx':
        return XlsxExcel(file)


class XlsxExcel:
    def __init__(self, file):
        self.wb = openpyxl.load_workbook(file, read_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.data = []

    def sheet_xuanze(self, n=0):
        self.ws = self.wb[self.wb.sheetnames[n]]
        return self.ws

    def datas(self):
        data = []
        for row in self.ws.values:
            print(list(row))
            data.append(list(row))
        print('数据读取完毕'.center(33, '-'))
        self.data = data[1:]
        return self.data

    def all_title(self):
        return self.ws.rows[0]

    def bufen_titles(self, lie_titiles=None):
        if not lie_titiles:
            return self.all_title()
        title = []
        for i in lie_titiles:
            if i in self.all_title():
                title.extend([i])
        return title

    def count(self):
        return len(self.ws.nrows) - 1

    def search_lies(self, lie_titles):
        row_datas = []
        for i in range(1, self.ws.nrows):
            row_data = []
            for lie_title in lie_titles:
                if lie_title in self.all_title():
                    row_data.extend(
                        [self.ws.row_values(rowx=i, start_colx=0, end_colx=None)[self.search_index(lie_title)]])
            row_datas.append(row_data)
        return row_datas
    def search_index(self, lie_title):
        if lie_title in self.all_title():
            return self.all_title().index(lie_title)
