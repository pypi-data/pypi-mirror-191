"""
Required: pip install openpyxl defusedxml
"""
from __future__ import annotations
import logging
from typing import Any
from datetime import datetime
from pathlib import Path

from . import filemgr
from .base import Format

try:
    from openpyxl import load_workbook, Workbook, DEFUSEDXML
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.worksheet.table import Table, TableColumn, TableFormula, TableStyleInfo
    from openpyxl.worksheet.formula import DataTableFormula, ArrayFormula
    from openpyxl.cell.cell import Cell
    from openpyxl.styles.differential import DifferentialStyle, DifferentialStyleList
    from openpyxl.styles.fills import PatternFill
    from openpyxl.utils import range_boundaries, get_column_letter

    _import_error = None

    logger = logging.getLogger(__name__)

    _cache = {
        'defusedxml_alert': False,
    }


    class ExcelFormat(Format):
        @classmethod
        def is_available(cls):
            return _import_error is None
        

        def __init__(self, **kwargs):
            super().__init__(**kwargs)

            if not self.is_available():
                raise ValueError(f"cannot use {self.__class__.__name__}: {_import_error}")

            if not DEFUSEDXML and not _cache['defusedxml_alert']:
                logger.warning("By default openpyxl does not guard against quadratic blowup or billion laughs xml attacks. To guard against these attacks install defusedxml.")
                _cache['defusedxml_alert'] = True


        def format_value(self, value: Any) -> str:
            if value is None:
                return None
            elif isinstance(value, datetime) and value.tzinfo:
                # Excel does not support timezones in datetimes
                if self.timezone:
                    value = value.astimezone(self.timezone)
                return value.replace(tzinfo=None)
            else:
                return super().format_value(value)


        def open_file(self):
            self.read_existing_file_if_any()

            # Open path for binary writting
            return filemgr.open_file(self.workbook_path, mode='wb', newline=self.newline, encoding=self.encoding, mkdir=True)


        workbook_path: str
        table_name: str
        workbook: Workbook
        worksheet: Worksheet
        previous_table: Table|None
        
        def read_existing_file_if_any(self):
            if hasattr(self, 'workbook_path'):
                return
            
            # Determine actual file path and table name
            self.workbook_path, self.table_name = self.flexout.split_path_target()
            if not self.table_name:
                self.table_name = self.DEFAULT_TABLE_NAME

            # Open workbook and search for existing table
            if filemgr.exists(self.workbook_path):
                with filemgr.open_file(self.workbook_path, 'rb') as fd:
                    self.workbook: Workbook = load_workbook(fd)

                self.previous_table = None
                for name in self.workbook.sheetnames:
                    self.worksheet: Worksheet = self.workbook[name]
                    if self.table_name in self.worksheet.tables:
                        self.previous_table = self.worksheet.tables[self.table_name]
                        break

                if not self.previous_table:
                    # table not found: we create a new worksheet
                    self.worksheet: Worksheet = self.workbook.create_sheet(title=self.table_name)

            else:
                self.workbook: Workbook = Workbook()
                self.worksheet: Worksheet = self.workbook.active
                self.worksheet.title = self.table_name
                self.previous_table = None


        headers_mapping: dict[int,int]|None
        unmanaged_col_indices: list[int]|None
        column_formats: dict[int,dict[str,Any]]|None
        previous_last_row_index: int|None
        first_col_index: int
        first_row_index: int
        last_col_index: int
        next_row_index: int
        warned_additional_columns: bool

        def append_headers(self, headers: list[str]):
            headers = super().append_headers(headers)

            self.headers_mapping = None
            self.unmanaged_col_indices = None
            self.column_formats = None

            self.read_existing_file_if_any()
            if self.previous_table:
                # Get existing table boundaries
                self.first_col_index, self.first_row_index, self.last_col_index, self.previous_last_row_index = range_boundaries(self.previous_table.ref)

                # Determines previous headers   
                previous_headers: list[str] = []
                column: TableColumn
                for i, column in enumerate(self.previous_table.tableColumns):
                    previous_headers.append(column.name)
                    self.prepare_column_format(col_index=self.first_col_index + i, column=column)

                # Determines headers mapping
                new_headers = []
                for pos, header in enumerate(headers):
                    try:
                        previous_pos = previous_headers.index(header)
                    except ValueError:
                        new_headers.append(header)
                        previous_pos = len(previous_headers) + len(new_headers) - 1

                    if previous_pos != pos:
                        if self.headers_mapping is None:
                            self.headers_mapping = {}
                        self.headers_mapping[pos] = previous_pos

                # Determines header columns that are not handled by us
                for pos, header in enumerate(previous_headers):
                    if not header in headers:
                        if self.unmanaged_col_indices is None:
                            self.unmanaged_col_indices = []
                        self.unmanaged_col_indices.append(self.first_col_index + pos)
                
                # Add new headers
                if new_headers:
                    logger.info(f"add header for table {self.table_name}: {', '.join(new_headers)} - was missing in existing table")

                    for header in new_headers:
                        self.last_col_index += 1
                        self.set_cell(self.first_row_index, self.last_col_index, header)

                        # Mark data as '?' (unknown) for the column if we append
                        if self.append:
                            r = self.first_row_index + 1
                            while r <= self.previous_last_row_index:
                                self.set_cell(r, self.last_col_index, '?')
                                r += 1

                # Determine next row
                if self.append:
                    self.next_row_index = self.previous_last_row_index + 1
                else:
                    self.next_row_index = self.first_row_index + 1
            
            else:
                # Table does not exist: write headers
                self.previous_last_row_index = None
                self.first_col_index = 1
                self.first_row_index = 1
                self.last_col_index = 0

                for header in headers:
                    self.last_col_index += 1
                    self.set_cell(self.first_col_index, self.last_col_index, header)

                self.next_row_index = self.first_row_index + 1

            self.warned_additional_columns = False

            return headers


        def append_row(self, row: list):
            if not hasattr(self, 'headers_mapping'):
                raise ValueError(f"headers not appended")

            row = super().append_row(row)

            # Write managed data
            for i in range(0, len(row)):
                if i >= len(self.headers):
                    if not self.warned_additional_columns:
                        logger.warning(f"ignore row values at index >= {len(self.headers)} (first occurence on row {len(self.rows)})")
                        self.warned_additional_columns = True
                else:
                    col_index = self.first_col_index + (self.headers_mapping.get(i, i) if self.headers_mapping else i)
                    cell = self.set_cell(self.next_row_index, col_index, row[i])
                    self.apply_column_format(cell)

            # Erase cell (except formula) and apply format on unmanaged data
            if self.unmanaged_col_indices:
                for col_index in self.unmanaged_col_indices:
                    cell = self.erase_cell(self.next_row_index, col_index, keep_formula=True)
                    self.apply_column_format(cell)

            self.next_row_index += 1
        
        
        def before_end(self):
            if not self.flexout.target_is_path:
                raise ValueError(f"can only use ExcelFormat with path target")

            if not self.headers:
                raise ValueError(f"cannot use ExcelFormat without headers")

            # Erase data outside of table
            if self.previous_last_row_index is not None:
                r = self.next_row_index
                while r <= self.previous_last_row_index:
                    c = self.first_col_index
                    while c <= self.last_col_index:
                        self.erase_cell(r, c)
                        c += 1
                    r += 1

            # Create table
            table_ref = f"{get_column_letter(self.first_col_index)}{self.first_row_index}:{get_column_letter(self.last_col_index)}{self.next_row_index-1}"
            if self.previous_table:
                if self.previous_table.ref != table_ref:
                    self.recreate_table(table_ref)
            else:
                table = Table(name=self.table_name, displayName=self.table_name, ref=table_ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                self.worksheet.add_table(table)

            # Save the file
            self.workbook.save(self.flexout.file)


        def recreate_table(self, new_ref):
            newcolumns = []

            for i in range(0, self.last_col_index - self.first_col_index + 1):
                name = self.worksheet.cell(self.first_row_index, self.first_col_index + i).value
                newcolumn = TableColumn(id=i+1, name=name)
                newcolumns.append(newcolumn)

                if i < len(self.previous_table.tableColumns):
                    prevcolumn: TableColumn = self.previous_table.tableColumns[i]
                    newcolumn.dataCellStyle = prevcolumn.dataCellStyle
                    newcolumn.dataDxfId = prevcolumn.dataDxfId # refers to workbook._differential_styles
                    newcolumn.calculatedColumnFormula = prevcolumn.calculatedColumnFormula

            newtable = Table(name=self.table_name, displayName=self.table_name, ref=new_ref, tableColumns=newcolumns, autoFilter=self.previous_table.autoFilter, sortState=self.previous_table.sortState)
            newtable.tableStyleInfo = self.previous_table.tableStyleInfo
            
            del self.worksheet.tables[self.table_name]
            self.worksheet.add_table(newtable)


        # -------------------------------------------------------------------------
        # Helpers
        # -------------------------------------------------------------------------
        DEFAULT_TABLE_NAME = 'Flexout'

        def set_cell(self, row_index: int, col_index: int, value) -> Cell:
            cell: Cell = self.worksheet.cell(row_index, col_index)

            try:
                cell.value = value
            except ValueError as err:
                if str(err).startswith('Cannot convert'):
                    cell.value = str(value)
                else:
                    raise

            return cell


        def erase_cell(self, row_index: int, col_index: int, keep_formula = False) -> Cell:
            cell: Cell = self.worksheet.cell(row_index, col_index)
            cell.style = 'Normal'
            
            if not (keep_formula and cell.data_type == 'f'):
                cell.value = None
            
            return cell


        def apply_column_format(self, cell: Cell|tuple(int,int)):
            if not self.column_formats:
                return

            if not isinstance(cell, Cell):
                row_index, col_index = cell
                cell = self.worksheet.cell(row_index, col_index)

            fmt = self.column_formats.get(cell.col_idx, None)
            if fmt is None:
                return

            if 'formula' in fmt:
                formula = fmt['formula']
                if isinstance(formula, ArrayFormula):
                    pass # TODO: not supported yet
                else:
                    cell.value = formula

            if 'style' in fmt:
                cell.style = fmt['style']

            for fmt_key, fmt_value in fmt.items():
                if fmt_key in ['formula', 'style']:
                    continue
                setattr(cell, fmt_key, fmt_value)


        def prepare_column_format(self, col_index: int, column: TableColumn) -> dict[str,Any]|None:
            if not self.column_formats:
                self.column_formats = {}

            fmt: dict[str,Any] = None

            # Read dataCellStyle
            if column.dataCellStyle:
                if fmt is None:
                    fmt = {}

                fmt['style'] = column.dataCellStyle
            
            # Read dxf
            if column.dataDxfId is not None:
                if fmt is None:
                    fmt = {}

                dxf: DifferentialStyle = self.workbook._differential_styles[column.dataDxfId]

                if dxf.numFmt:
                    fmt['number_format'] = dxf.numFmt.formatCode
                else:
                    if not 'style' in fmt:
                        fmt['number_format'] = self._DEFAULT_NUMBER_FORMAT

                fmt['alignment'] = dxf.alignment if dxf.alignment else self._DEFAULT_ALIGNMENT
                fmt['border'] = dxf.border if dxf.border else self._DEFAULT_BORDER
                fmt['font'] = dxf.font if dxf.font else self._DEFAULT_FONT
                fmt['protection'] = dxf.protection if dxf.protection else self._DEFAULT_PROTECTION
                fmt['fill'] = PatternFill(fill_type=dxf.fill.fill_type, bgColor=dxf.fill.fgColor, fgColor=dxf.fill.bgColor) if dxf.fill else self._DEFAULT_FILL # NOTE: fgcolor and bgcolor are inversed in DifferentialStyle

            # Read formula
            if column.calculatedColumnFormula:
                if fmt is None:
                    fmt = {}

                formula = column.calculatedColumnFormula
                if formula.array:
                    fmt['formula'] = ArrayFormula(formula.attr_text)
                else:
                    fmt['formula'] = '=' + formula.attr_text
            
            # Register format
            if fmt is not None:
                self.column_formats[col_index] = fmt
        

        _DEFAULT_NUMBER_FORMAT = 'General'

        _DEFAULT_FILL = PatternFill(fill_type=None)

        _DEFAULT_ALIGNMENT = None # openpyxl.styles.alignment.Alignment
        _DEFAULT_BORDER = None # openpyxl.styles.alignment.Border
        _DEFAULT_FONT = None # openpyxl.styles.fonts.Font
        _DEFAULT_PROTECTION = None # openpyxl.styles.protection.Protection


# -----------------------------------------------------------------------------
except ImportError as err:
    _import_error = str(err)

    class ExcelFormat(Format):
        @classmethod
        def is_available(cls):
            return False
        

        def __init__(self, **kwargs):
            super().__init__(**kwargs)

            if not self.is_available():
                raise ValueError(f"cannot use {self.__class__.__name__}: {_import_error}")
